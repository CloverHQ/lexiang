#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2022-10-30 20:30
# @Author  : liuwenchao
# @File    : record
# @Software: IntelliJ IDEA
import json
import os
import random
import re
import sys
import time

import openpyxl
from tqdm import tqdm


def get_data_batch(file_name):
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    # 获取指定的sheet页对象
    sheet = workbook["Sheet1"]
    rows = sheet.rows
    list = []
    for row in rows:
        record = {
            "carNo": None,
            "date": None,
            "addr": None
        }

        for data in row[1:11]:
            if data.row < 4 or data.column > 11:
                continue
            if data.column == 2:
                record["date"] = str(data.value)
                continue
            if data.column == 4:
                record["carNo"] = data.value
                continue
            if data.column == 5 and record["carNo"] is None:
                record["carNo"] = data.value
                continue
            if data.column == 11:
                record["addr"] = data.value
                continue
        if record["carNo"] is not None:
            list.append(record)
    return list


def created_excel(header, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    # 添加表头（不需要表头可以不用加）
    data.insert(0, list(header))
    # 开始遍历数组
    for row_index, row_item in enumerate(data):

        for col_index, col_item in enumerate(row_item):
            # 写入
            sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)

    # 写入excel文件 如果path路径的文件不存在那么就会自动创建
    workbook.save('记录.xlsx')


reg = '^[\u4E00-\u9FA5\uF900-\uFA2D\u0020]*$'

if __name__ == "__main__":
    str1 = '██    ██  █████   ██████'
    str2 = ' ██  ██  ██   ██ ██    ██'
    str3 = '  ████   ███████ ██    ██'
    str4 = '   ██    ██   ██ ██    ██'
    str5 = '   ██    ██   ██  ██████'
    # 获取数据
    print(str1)
    print(str2)
    print(str3)
    print(str4)
    print(str5)

    f_list = []

    print('正在检查运行环境....')

    if not os.path.exists('user_info.json'):
        input('user_info文件不存在，请检查配置文件是否存在，按回车退出。')
        sys.exit()

    for file_name in os.listdir(os.getcwd()):
        if os.path.splitext(file_name)[1] == '.xlsx':
            f_list.append(file_name)

    if len(f_list) == 0:
        input('当前文件夹没有要解析的文件, 请检查文件目录是否正确, 按回车键退出..')
        sys.exit()
    time.sleep(2)

    print('运行环境检查通过.... ^_^')
    print('------------------------')
    print('文件列表：')
    for idx, f in enumerate(f_list):
        print('%d. %s' % (idx, f))
    print('------------------------')

    if len(f_list) > 1:
        while True:
            index = int(input('请输入要解析的文件序号... \r\n').strip())
            if index >= len(f_list):
                print('文件索引不存在, 请重新输入...')
            else:
                file_name = f_list[index]
                break
    else:
        file_name = f_list[0]
    before = time.time()
    print('正在解析请稍后....')
    records = get_data_batch(file_name)
    print("解析完成，正在生成新的Excel数据表...")
    header = ['日期/出站时间', '车号', '核定载质量', '出厂站核载质量', '驾驶员姓名', '联系方式', '送达/采购地址']

    data = []

    with open("user_info.json", encoding='utf-8') as fw:
        infos = json.load(fw)
        for record in tqdm(records, desc="生成进度"):
            for info in infos:
                if info['carNo'].strip() == record['carNo'].strip():
                    mass = int(info['ratedLoadingMass'][:-2])

                    result = [record['date'], info['carNo'], info['ratedLoadingMass'],
                              str(round(random.randint(mass - 3000, mass - 1000), -1)) + 'kg', info['name'],
                              info['contact'],
                              record['addr']]
                    data.append(result)
        not_exists = filter(lambda x: not re.search(reg, x), set(map(lambda x: x['carNo'].strip(), records)) - set(
            map(lambda x: x['carNo'].strip(), infos)))

        created_excel(header, data)
        fileName = '未收集人员信息的车牌号.txt'
        with open(fileName, 'w', encoding='utf-8')as file:
            file.writelines([line + '\n' for line in not_exists])
        input('[Mission Accomplished, Press Enter to exit...] 用时: %s 感谢使用！bye///' % str(time.time() - before))
